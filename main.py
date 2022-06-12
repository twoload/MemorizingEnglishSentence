import pandas as pd
from openpyxl import load_workbook
from fpdf import FPDF
#from excel import Excel

class PDF(FPDF):
    def header(self):
        # Arial bold 15
        #self.set_font('Arial', 'B', 15)
        # Move to the right
        #self.cell(80)
        # Title
        #self.cell(30, 10, title, 1, 0, 'C')
        # Line break
        self.ln(20)

    # Page footer
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        # Arial italic 8
        self.set_font('Arial', 'I', 8)
        # Page number
        self.cell(0, 10, 'Page ' + str(self.page_no()) + '/{nb}', 0, 0, 'C')


def isWorkSheet(excelname, worksheet):
    wb = load_workbook(excelname, read_only=True)
    if worksheet in wb.sheetnames:
        #print(f'exist {worksheet}')
        return True
    else:
        #print(f'no exist {worksheet}')
        return False

def inputSubTitlePage(index):
    if isWorkSheet('mainPage.xlsx', str(index)):
        print(f'there is worksheet {str(index)}')
        subPage = pd.read_excel('mainPage.xlsx', str(index))
    else:
        print(f'there is no worksheet {str(index)}')
        return

    print(subPage)
    while 1:
        print("==========================")
        print(f'Edit Contents ({str(index)})')
        print("==========================")
        print("1.add")
        print("2.remove")
        print("3.display")
        print("4.edit contents")
        print("5.save to excel")
        print("0.exit")
        print("==========================")
        print("choose>")
        choose = int(input())

        if choose == 1: #add
            print('input English >')
            english = input()
            print('input Korean >')
            korean = input()
            subPage.loc[len(subPage.index)] = [english, korean]
            print(subPage)
        elif choose == 2: #remove
            print("choose index to remove>")
            index = int(input())
            subPage = subPage.drop(subPage.index[index])
            print(subPage)
        elif choose == 3: #display
            print(subPage)
        elif choose == 4: #edit
            print(subPage)
            print("edit index to edit>")
            index = int(input())
            print(subPage.loc[[index]])
            print("Do you want to change English? > (y/n)")
            ans = input()
            if ans == 'y':
                print('input English>')
                english = input()
            print("Do you want to change Korean? > (y/n)")
            ans = input()
            if ans == 'y':
                print('input Korean>')
                korean = input()
            subPage.loc[[index]] = [english, korean]
            print(subPage)
        elif choose == 5: #save to excel
            subPage.to_excel('./mainPage.xlsx', sheet_name=str(index))
        elif choose == 0:
            break
        else:
            print('wrong input')


def inputMainPage():
    mainPage = pd.read_excel('mainPage.xlsx', 'mainPage')
    print(mainPage)

    choose = 0
    while 1:
        print("==========================")
        print("Edit Title")
        print("==========================")
        print("1.add")
        print("2.remove")
        print("3.display")
        print("4.edit sub title")
        print("5.save to excel")
        print("0.exit")
        print("==========================")
        print("choose>")
        choose = int(input())

        if choose == 1:
            print("input date(ex.220501)")
            date = input()
            print("input title(ex.Oxford Reading Tree)")
            title = input()
            print("input subTitle(ex.book1)")
            subTitle = input()
            mainPage.loc[len(mainPage.index)] = [date, title, subTitle]
            print(mainPage)

        elif choose == 2:
            print("choose index to remove>")
            index = int(input())
            mainPage = mainPage.drop(mainPage.index[index])
            print(mainPage)
        elif choose == 3:
            print(mainPage)
        elif choose == 4:
            print(mainPage)
            print('choose index>')
            index = int(input())
            inputSubTitlePage(index)
        elif choose == 5:
            mainPage.to_excel('./mainPage.xlsx', sheet_name='mainPage')
            pass
        elif choose == 0:
            break
        else:
            print("wrong input")
    return

def study():
    mainPage = pd.read_excel('mainPage.xlsx', 'mainPage')

    while 1:
        studyMenu = 0
        print("==========================")
        print(" study Menu ")
        print("==========================")
        print("1.create pdf")
        print("0.exit")
        print("==========================")
        print("choose number > ")
        studyMenu = int(input())

        if studyMenu == int(1):
            print(mainPage)
            print('select index >')
            index = int(input())
            if isWorkSheet('mainPage.xlsx', str(index)):
                subPage = pd.read_excel('mainPage.xlsx', str(index))
            print(subPage)

            nameTitle = mainPage['title'][index] + '___' + mainPage['subTitle'][index]
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('sysfont', '', 'NEXONLv1GothicRegular.ttf')
            pdf.set_font("sysfont", size=10)
            line_height = pdf.font_size * 2.5

            # head
            pdf.set_font(style = 'U')
            pdf.cell(w = 0, h = 10, txt = nameTitle, border = 0, align = 'C')
            pdf.set_font(style ='')
            pdf.ln(line_height*2)
            # contents
            col_width = pdf.epw / 2  # distribute content evenly
            for index in subPage.index:
                eng = subPage['English'][index]
                kor = subPage['Korean'][index]
                data_row = (eng, kor)
                for datum in data_row:
                    pdf.multi_cell(col_width, line_height, datum, border=0, new_x="RIGHT", new_y="TOP", max_line_height=pdf.font_size)
                pdf.ln(line_height)

            pdf.output(f'{nameTitle}.pdf')
        elif studyMenu == int(0):
            break
        else:
            print('wrong input')


'''
            # create pdf
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('Arial', '', 16)
            pdf = FPDF('P', 'mm', 'A4')
            #pdf.title(mainPage['title'][index], mainPage['subTitle'][index])
            nameTitle = mainPage['title'][index] + mainPage['subTitle'][index]
            pdf.set_title('nameTitle')
            pdf.cell(40, 10, 'hello')
            #for index in subPage.index:
                #print(subPage['English'][index], subPage['Korean'][index])
            #    pdf.cell(40, 10, subPage['English'][index])
            #    pdf.cell()
            pdf.output('output.pdf', 'F')
    '''

def showWelcome():
    print("==========================")
    print(' ~ Welcome to M.E.S ~ ')
    #print("==========================")

def showMainMenu():
    while 1:
        numMainMenu = 0
        print("==========================")
        print(" Main Menu ")
        print("==========================")
        print("1.input English sentences")
        print("2.upload answers (excel files)")
        print("3.study")
        print("0.exit")
        print("==========================")
        print("choose number > ")
        numMainMenu = int(input())

        if numMainMenu == int(1):
            inputMainPage()
        elif numMainMenu == int(2):
            pass
        elif numMainMenu == int(3):
            study()
        elif numMainMenu == int(0):
            break
        else:
            print("wrong number")

    return


def main():
    showWelcome()
    showMainMenu()

if __name__ == "__main__":
    main()