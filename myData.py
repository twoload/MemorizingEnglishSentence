import pandas as pd
from myExcel import *
import openpyxl
import os.path

MAINPAGE_FILENAME = './mainPage.xlsx'
MAINPAGE_SHEETNAME = 'mainPage'

class mainPageDatabase:

    def __init__(self):
        self.df = pd.DataFrame([], columns=['date', 'title', 'sub'])
        self.init()

    def init(self):
        # file check
        if not os.path.exists(MAINPAGE_FILENAME):
            self.create()
            return True
        else:
            read_df = self.load()
            if read_df.empty:
                #print('read_df empty')
                self.create()
                return True
            else:
                #print('copy read_df into self.df')
                num_matched_column = self.df.columns.intersection(read_df.columns).size
                if num_matched_column == len(self.df.columns):
                    self.df = read_df
                    #print(self.df)
                    return True
                else:
                    print('not match column head')
                    return False

    def create(self):
        print('mainPageDatabase create')
        dataframe2excel(self.df, MAINPAGE_FILENAME, MAINPAGE_SHEETNAME)

    def load(self):
        print('mainPageDatabase load: %s' % MAINPAGE_FILENAME)
        ret = excel2dataframe(MAINPAGE_FILENAME, MAINPAGE_SHEETNAME)
        return ret

    def print_data_lists_by_row(self):
        for row in range(len(self.df)):
            row_list = []
            for col in range(len(self.df.columns)):
                row_list.append(self.df.iloc[row][col])
            print(row_list)

    def get_data_frame(self):
        return self.df

    def add_empty_row(self, item):
        self.df.loc[len(self.df.index)] = []

    def remove_row(self, row_index):
        self.df = self.df.drop(self.df.index[row_index])

    def update_item(self, row_index, col_name, value):
        self.df.loc[row_index, [col_name]] = [value]
        #print(self.df)

    def save(self):
        dataframe2excel(self.df, MAINPAGE_FILENAME, MAINPAGE_SHEETNAME)

class subPageDatabase:

    def __init__(self):
        self.df = pd.DataFrame([], columns=['English', 'Korean'])
        self.dataframes = dict()
        self.init()

    def init(self):
        #print(xl.sheet_names)
        # check file
        if not os.path.exists(MAINPAGE_FILENAME):
            print('subPageDatabase:error: file not exists (%s)' % MAINPAGE_FILENAME)
            return
        # check sheet
        workbook = openpyxl.load_workbook(MAINPAGE_FILENAME)

        # load sheet info
        for sheet in workbook.sheetnames:
            self.init_by_sheetname(sheet)

        #print(self.dataframes)

    def init_by_sheetname(self, sheetname):
        read_df = self.load(sheetname)
        if read_df.empty:
            #print('read_df empty')
            self.create_sheet(sheetname)
            return True
        else:
            #print('copy read_df into self.df')
            num_matched_column = self.df.columns.intersection(read_df.columns).size
            if num_matched_column == len(self.df.columns):
                self.update_sheet(sheetname, read_df)
                return True
            else:
                print('not match column head')
                return False

    def create_sheet(self, sheetname):
        print('subPageDatabase %s (%s) create' %(MAINPAGE_FILENAME, sheetname))
        df = pd.DataFrame([], columns=['English', 'Korean'])
        # save into dataframe
        self.dataframes[sheetname] = df
        # save into excel
        self.save(sheetname)

    def update_sheet(self, sheetname, df):
        print('subPageDatabase %s (%s) update' % (MAINPAGE_FILENAME, sheetname))
        # save into dataframe
        self.dataframes[sheetname] = df
        # save into excel
        self.save(sheetname)

    def remove_sheet(self, sheetname):
        del self.dataframes[sheetname]
        workbook = openpyxl.load_workbook(MAINPAGE_FILENAME)
        del workbook[sheetname]
        workbook.save(MAINPAGE_FILENAME)

    def show(self):
        print(self.dataframes)

    def save(self, sheetname):
        print('subPageDatabase:save: sheet_name = %s ' % sheetname)
        df = pd.DataFrame([], columns=['English', 'Korean'])
        if self.isSheetExist(sheetname):
            df = self.dataframes[sheetname]

        with pd.ExcelWriter(MAINPAGE_FILENAME, mode='a', engine='openpyxl', if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheetname, index=False)

    def add_empty_row(self, sheet, item):
        self.dataframes[sheet].loc[len(self.dataframes[sheet].index)] = []

    def remove_row(self, sheet, row_index):
        self.dataframes[sheet] = self.dataframes[sheet].drop(self.dataframes[sheet].index[row_index])

    def update_item(self, sheet, row_index, col_name, value):
        self.dataframes[sheet].loc[row_index, [col_name]] = [value]

    def load(self, sheetname):
        print('subPageDatabase load: %s (%s)' %(MAINPAGE_FILENAME, sheetname))
        df = excel2dataframe(MAINPAGE_FILENAME, sheetname)
        return df
        #print(self.df)

    def loads(self):
        sheet_names = self.sheetnames()
        for sheetname in sheet_names:
            if sheetname is not MAINPAGE_SHEETNAME:
                if 1 <= int(sheetname) <= 100:
                    df = self.load(sheetname)
                    self.dataframes[sheetname] = df
        #print(self.dataframes)

    def sheetnames(self):
        xls = pd.ExcelFile(MAINPAGE_FILENAME, engine="openpyxl")
        sheet_list = xls.sheet_names
        return sheet_list

    def isSheetExist(self, sheetname):
        sheet_list = self.sheetnames()
        if sheetname in sheet_list:
            return True
        else:
            return False

    def printDataListsByRow(self, sheetname):
        df = self.dataframes[sheetname]
        for row in range(len(df)): #row
            row_list = []
            for col in range(len(df.columns)): #col
                row_list.append(df.iloc[row][col])
            print(row_list)

    def setCurrentDataFrame(self, sheetname):
        self.df = self.dataframes[sheetname]

    def getCurrentDataFrame(self, sheetname):
        return self.df
'''
db = mainPageDatabase()
db.init()
db.update_item(1, 'Date', 220701)
'''
db_sub = subPageDatabase()
db_sub.init()
#db_sub.create('1')
#db_sub.create('1')
#
#db_sub.loads()
#db_sub.create_sheet('3')
#db_sub.create_sheet('11')
#db_sub.create_sheet('12')
#db_sub.create_sheet('13')
#db_sub.remove_sheet('11')
#db_sub.sheetnames()