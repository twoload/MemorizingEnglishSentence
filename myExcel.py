import sys
import pandas as pd
from openpyxl import load_workbook

def excel2dataframe(excelFilename, worksheetName):
    df = pd.read_excel(excelFilename, worksheetName)
    print(df)
    return df

def dataframe2excel(df, excelFilename, worksheetName):
    df.to_excel(excelFilename, sheet_name = worksheetName)


def isWorkSheet(excelFilename, worksheet):
    wb = load_workbook(excelFilename, read_only=True)
    if worksheet in wb.sheetnames:
        #print(f'exist {worksheet}')
        return True
    else:
        #print(f'no exist {worksheet}')
        return False